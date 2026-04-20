import os
import io
import shutil
from openpyxl import load_workbook
from datetime import datetime
import re

from flask import Flask, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

### 서열 정보 ###
# 1. 엑셀 파일 불러오기(wb: workbook)
def transform_excel_to_wipo_xml(excel_file):

    file_stream = io.BytesIO(excel_file.read())

    wb = load_workbook(file_stream, data_only = True)
    ws_sequence = wb['서열정보']

    # 서열정보 시트 선택
    wb.active = ws_sequence
    count = ws_sequence.max_row - 1

    SequenceDataTemp = """
    \t<SequenceData sequenceIDNumber="{n}">
    \t\t<INSDSeq>
    \t\t\t<INSDSeq_length>{sequence_len}</INSDSeq_length>
    \t\t\t<INSDSeq_moltype>{moltype}</INSDSeq_moltype>
    \t\t\t<INSDSeq_division>PAT</INSDSeq_division>
    \t\t\t<INSDSeq_feature-table>
    \t\t\t\t<INSDFeature>
    \t\t\t\t\t<INSDFeature_key>source</INSDFeature_key>
    \t\t\t\t\t<INSDFeature_location>1..{sequence_len}</INSDFeature_location>
    \t\t\t\t\t<INSDFeature_quals>
    \t\t\t\t\t\t<INSDQualifier>
    \t\t\t\t\t\t\t<INSDQualifier_name>mol_type</INSDQualifier_name>
    \t\t\t\t\t\t\t<INSDQualifier_value>{details}</INSDQualifier_value>
    \t\t\t\t\t\t</INSDQualifier>
    \t\t\t\t\t\t<INSDQualifier id="q{id}">
    \t\t\t\t\t\t\t<INSDQualifier_name>organism</INSDQualifier_name>
    \t\t\t\t\t\t\t<INSDQualifier_value>{organism}</INSDQualifier_value>
    \t\t\t\t\t\t</INSDQualifier>
    \t\t\t\t\t</INSDFeature_quals>
    \t\t\t\t</INSDFeature>
    \t\t\t</INSDSeq_feature-table>
    \t\t\t<INSDSeq_sequence>{sequence}</INSDSeq_sequence>
    \t\t</INSDSeq>
    \t</SequenceData>"""

    # 시퀀스 데이터 누적용 변수
    sequence_data = ""
    for row in range(2, ws_sequence.max_row + 1):
        id = row
        moltype = ws_sequence.cell(row=row, column=2).value
        details = ws_sequence.cell(row=row, column=3).value
        sequence = ws_sequence.cell(row=row, column=4).value.replace(" ","")
        if moltype == "AA":
            sequence = sequence.upper()
        else:
            sequence = sequence.lower()       
        sequence_len = len(sequence)
        organism = ws_sequence.cell(row=row, column=5).value
        temp = SequenceDataTemp.format(n=str(row-1), id=id, moltype=moltype, details=details, sequence=sequence, sequence_len=sequence_len, organism=organism)
        sequence_data += temp

    production_date = datetime.now().strftime('%Y-%m-%d')

    xml_basic_template = """<?xml version="1.0" encoding="UTF-8"?>
    <!DOCTYPE ST26SequenceListing PUBLIC "-//WIPO//DTD Sequence Listing 1.3//EN" "ST26SequenceListing_V1_3.dtd">
    <ST26SequenceListing   dtdVersion="V1_3" fileName=".xml" softwareName="WIPO Sequence" softwareVersion="2.3.0" productionDate="{production_date}">
    \t<ApplicantFileReference></ApplicantFileReference>
    \t<ApplicantName languageCode="ko"></ApplicantName>
    \t<ApplicantNameLatin></ApplicantNameLatin>
    \t<InventionTitle languageCode="ko"></InventionTitle>
    \t<InventionTitle languageCode="en"></InventionTitle>
    \t<SequenceTotalQuantity>{count}</SequenceTotalQuantity>{sequence_data}
    </ST26SequenceListing>
    """
    
    final_xml = xml_basic_template.format(
        production_date = production_date,
        count = count,
        sequence_data = sequence_data
    )

    return final_xml

@app.route('/upload', methods['POST'])
def handle_upload():
    if 'excel_file' not in request.files:
        return "파일이 전송되지 않았습니다.", 400

    excel_file = request.files['excel_file']

    try:
        result_xml = transform_excel_to_wipo_xml(excel_file)

        return send_file(
            io.BytesIO(result_xml.encode('utf-8')),
            mimetype='application/xml',
            as_attachment=True,
            download_name='WIPO_Sequence_Result.xml'
        )
    
    except Exception as e:
        return f"변환 중 오류 발생: {str(e)}", 500
    
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host='0.0.0.0', port=port)